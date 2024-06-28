import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from copy import copy
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side
import os
from database import (get_database_connection, get_database_engine, get_database_engine_e_eiis,
                      get_database_connection_e_eiis)


def location_period(year, month):
    pivot_df_list: list = []
    location_id_list: list = []
    ReqNo_code_list: list = []
    location_name_list: list = []
    status = None

    with get_database_connection_e_eiis() as conn:
        try:
            # Create a cursor object
            cursor = conn.cursor()

            # SQL query with placeholders for YEAR and MONTH
            query = """
                    SELECT head.REQ_NO , head.LOCATION_ID FROM req_head AS head WHERE 
                    YEAR(head.PERIOD) = %s AND MONTH(head.PERIOD) = %s;
                    """

            # Execute the query with dynamic parameters
            cursor.execute(query, (year, month))

            # Fetch all records
            records = cursor.fetchall()
        except Exception as error:
            print('The cause of error -->', error)
            status = "failed"
            return status, pivot_df_list, location_id_list, ReqNo_code_list

        try:
            # Get the database engine
            engine = get_database_engine_e_eiis()

            # SQL Query with parameters
            sql_query = """
                        SELECT head.REQ_NO, head.LOCATION_ID, location.Location_Name, detail.REQUEST_DATE, 
                        detail.ITEM_ID, item.Item_Name, detail.PACKAGE_ID, detail.QTY
                        FROM req_head AS head
                        JOIN req_detail AS detail ON detail.REQ_NO = head.REQ_NO
                        JOIN item ON item.Item_ID = detail.ITEM_ID
                        INNER JOIN location ON location.Location_ID = head.LOCATION_ID
                        WHERE head.REQ_NO = %s AND head.LOCATION_ID = %s;
                        """

            # Process each tuple
            for req_tuple in records:
                req_no, location_id = req_tuple  # Unpack the tuple
                df = pd.read_sql_query(sql_query, engine, params=(req_no, location_id))

                # Pivoting the DataFrame
                pivot_df = df.pivot_table(index=['REQ_NO', 'LOCATION_ID', 'Location_Name', 'ITEM_ID',
                                                 'Item_Name', 'PACKAGE_ID'],
                                          columns='REQUEST_DATE',
                                          values='QTY',
                                          aggfunc='sum',
                                          fill_value=0)

                # Resetting index to flatten the DataFrame
                pivot_df = pivot_df.reset_index()
                # Calculating sum of each row after the PackageId column
                pivot_df['Total Qty'] = pivot_df.iloc[:, 6:].sum(axis=1)
                Req_No = pivot_df['REQ_NO'].tolist()
                location_id = pivot_df['LOCATION_ID'].tolist()
                pivot_df.drop(columns=['LOCATION_ID'], inplace=True)

                location_name = pivot_df['Location_Name'].tolist()
                pivot_df.drop(columns=['Location_Name'], inplace=True)

                location_name_list.append(location_name[0])
                location_id_list.append(location_id[0])
                pivot_df_list.append(pivot_df)
                ReqNo_code_list.append(Req_No[0])
                status = "success"

        except Exception as error:
            print('The Cause of error -->', error)
            status = "failed"
        return status, pivot_df_list, location_id_list, ReqNo_code_list, location_name_list


def fetch_location_period_new_selected(year, month, sql_query_type):
    pivot_df_list = []
    location_id_list = []
    ReqNo_code_list = []
    location_name_list = []
    status = None

    with get_database_connection_e_eiis() as conn:
        try:
            # Create a cursor object
            cursor = conn.cursor()

            # SQL query with placeholders for YEAR and MONTH
            query = """
                    SELECT head.REQ_NO, head.LOCATION_ID 
                    FROM req_head_new AS head 
                    WHERE YEAR(head.PERIOD) = %s AND MONTH(head.PERIOD) = %s;
                    """

            # Execute the query with dynamic parameters
            cursor.execute(query, (year, month))

            # Fetch all records
            records = cursor.fetchall()
            print(records)
        except Exception as error:
            print('The cause of error -->', error)
            status = "failed"
            return status, pivot_df_list, location_id_list, ReqNo_code_list, location_name_list
        finally:
            cursor.close()  # Ensure the cursor is closed

    # Determine the SQL query based on sql_query_type
    if sql_query_type == "1":
        sql_query = """
                    SELECT head.REQ_NO, 
                           head.LOCATION_ID, 
                           location.Location_Name, 
                           detail.REQUEST_DATE, 
                           detail.ITEM_ID, 
                           item.Item_Name, 
                           detail.PACKAGE_ID, 
                           detail.QTY 
                    FROM req_head_new AS head 
                    JOIN req_detail_new AS detail 
                        ON detail.REQ_NO = head.REQ_NO
                    JOIN item 
                        ON item.Item_ID = detail.ITEM_ID
                    JOIN location 
                        ON location.Location_ID = head.LOCATION_ID
                    WHERE head.REQ_NO = %s 
                      AND head.LOCATION_ID = %s 
                      AND detail.SUPPLIER_ID != '';
                    """
    else:
        sql_query = """
                    SELECT head.REQ_NO, 
                           head.LOCATION_ID, 
                           location.Location_Name, 
                           detail.REQUEST_DATE, 
                           detail.ITEM_ID, 
                           item.Item_Name, 
                           detail.PACKAGE_ID, 
                           detail.QTY 
                    FROM req_head_new AS head 
                    JOIN req_detail_new AS detail 
                        ON detail.REQ_NO = head.REQ_NO
                    JOIN item 
                        ON item.Item_ID = detail.ITEM_ID
                    JOIN location 
                        ON location.Location_ID = head.LOCATION_ID
                    WHERE head.REQ_NO = %s 
                      AND head.LOCATION_ID = %s 
                      AND detail.SUPPLIER_ID = '';
                    """

    try:
        # Get the database engine
        engine = get_database_engine_e_eiis()

        # Process each tuple from the records
        for req_no, location_id in records:
            df = pd.read_sql_query(sql_query, engine, params=(req_no, location_id))
            if not df.empty:
                print(df)
                # Pivoting the DataFrame
                pivot_df = df.pivot_table(
                    index=['REQ_NO', 'LOCATION_ID', 'Location_Name', 'ITEM_ID', 'Item_Name', 'PACKAGE_ID'],
                    columns='REQUEST_DATE',
                    values='QTY',
                    aggfunc='sum',
                    fill_value=0
                ).reset_index()

                # Calculate the total quantity
                pivot_df['Total Qty'] = pivot_df.iloc[:, 6:].sum(axis=1)

                Req_No = pivot_df['REQ_NO'].tolist()
                loc_id = pivot_df['LOCATION_ID'].tolist()
                location_name = pivot_df['Location_Name'].tolist()

                # Drop the LOCATION_ID and Location_Name columns
                pivot_df.drop(columns=['LOCATION_ID', 'Location_Name'], inplace=True)

                # Append the results to the lists
                location_name_list.append(location_name[0])
                location_id_list.append(loc_id[0])
                pivot_df_list.append(pivot_df)
                ReqNo_code_list.append(Req_No[0])
            status = "success"

    except Exception as error:
        print('The cause of error -->', error)
        status = "failed"

    return status, pivot_df_list, location_id_list, ReqNo_code_list, location_name_list


def create_merge_excel(location_id_list, pivot_df_list, ReqNo_code_list, location_name_list):
    file_name = None
    file_path = None
    try:
        # Assuming location_name and month_variable are defined outside the function
        month_variable = "Apr-2024"

        # Create a new Excel workbook
        workbook = Workbook()
        workbook.remove(workbook["Sheet"])

        excel_file_path = []

        for pivot_df, location_code_variable, req_num, loc_name in zip(pivot_df_list, location_id_list, ReqNo_code_list,
                                                                       location_name_list):
            # Create a new sheet for each iteration
            sheet = workbook.create_sheet(title=f"{req_num}_{location_code_variable}")

            # Write DataFrame to Excel file starting from 6th row, first column
            rows = dataframe_to_rows(pivot_df, index=False, header=True)
            for r_idx, row in enumerate(rows, 6):
                for c_idx, value in enumerate(row, 1):
                    cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 6:  # If it's the first row, make it bold and center aligned
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    # Adjust column width based on length of value inside the cell
                    column_letter = chr(64 + c_idx)  # Convert column index to letter
                    max_length = max(len(str(cell.value)) for cell in sheet[column_letter] if
                                     cell.value)  # Get max length of values in column
                    adjusted_width = (max_length + 2) * 1.2  # Adjusting width for some padding
                    sheet.column_dimensions[column_letter].width = adjusted_width

            # Apply borders to the table
            max_row = sheet.max_row
            max_col = sheet.max_column
            border_style = Side(style='thin', color='000000')
            border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)

            for row in sheet.iter_rows(min_row=6, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.border = border

            # Calculate the width of the box based on the last column of the table
            last_column_width = sheet.column_dimensions[chr(64 + max_col)].width
            box_width = last_column_width + 2  # Adjusting width for some padding

            # Draw a box above the table from 5th row to 2nd row with the calculated width
            box_end_column = chr(64 + max_col)  # End column is the last column of the table
            box_range = sheet[f'A2:{box_end_column}5']  # Specify the range for the box
            for row in box_range:
                for cell in row:
                    if cell.row == 2 or cell.row == 5:  # Draw top and bottom borders
                        if cell.column == 1 or cell.column == max_col:  # Draw borders for the first and last columns
                            cell.border = Border(top=border_style, bottom=border_style, left=border_style,
                                                 right=border_style)
                        elif 1 < cell.column < max_col:  # Draw borders for cells in between the first and last columns
                            cell.border = Border(top=border_style, bottom=border_style)
                    elif 2 < cell.row < 5:  # Draw borders for cells in between the top and bottom rows
                        if cell.column == 1 or cell.column == max_col:  # Draw borders for the first and last columns
                            cell.border = Border(left=border_style, right=border_style)
                        elif 1 < cell.column < max_col:  # Draw borders for cells in between the first and last columns
                            pass

            # Merge cells and write text "Location Monthly Request" in the second row
            sheet.merge_cells(f'A2:{box_end_column}2')
            cell = sheet['A2']
            cell.value = "Location Monthly Request"
            cell.font = Font(bold=True, size=14)  # Increase the font size to 14
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Increase the height of the second row
            sheet.row_dimensions[2].height = 30  # Adjust the height as per your requirement

            # Merge cells in the 3rd row and add "Location Name :-" and location name
            sheet.merge_cells('C3:C3')
            cell = sheet['C3']
            cell.value = "Location Name :"
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='right')  # Align text to the right

            sheet['D3'] = loc_name  # Add the dynamic location name to the next cell
            sheet['D3'].alignment = Alignment(horizontal='left')  # Align text to the left

            # Find the column letter of the last column in the table
            last_column_letter = chr(64 + max_col - 1)

            # Add "Month -" in the last cell of the 3rd row
            month_cell = sheet[last_column_letter + '3']
            month_cell.value = "Month :"
            month_cell.font = Font(bold=True)

            # Find the column letter of the cell after the last column in the table
            month_value_cell_column = chr(64 + max_col)  # This assumes the box ends where the table ends

            # Add the month variable in the next cell after "Month -" in the 3rd row
            month_value_cell = sheet[month_value_cell_column + '3']
            month_value_cell.value = month_variable

            # Add "Location Code :" in the first cell of the 3rd column (beginning of the box)
            location_code_cell = sheet['A3']
            location_code_cell.value = "Location Code :"
            location_code_cell.font = Font(bold=True)
            # Adjust the width of column A to fit the text
            sheet.column_dimensions['A'].width = len("Location Code :") + 1  # Adjust the width as needed

            # Find the column letter of the cell after "Location Code :"
            location_code_value_cell_column = chr(65)  # Column B

            # Add the location code in the next cell after "Location Code :" in the 3rd row
            location_code_value_cell = sheet['B3']
            location_code_value_cell.value = location_code_variable

            # Append the file path to the list
            excel_file_path.append(f"{req_num}_{location_code_variable}")

        # Specify the folder path
        folder_path = r'C:\Users\Administrator\Downloads\eiis\location_request\merged_excel'

        # Ensure the folder exists, if not, create it
        os.makedirs(folder_path, exist_ok=True)

        current_datetime = datetime.datetime.now()
        file_name = f"Merged_Excel_{current_datetime.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        file_path = os.path.join(folder_path, file_name)

        # Save the workbook in the specified folder
        workbook.save(filename=file_path)
        excel_file_path.append(file_path)
        print(file_path)
        status = "success"

    except Exception as error:
        print('The cause of error -->', error)
        status = "failed"

    return status, file_path, file_name
