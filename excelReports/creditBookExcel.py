from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
from datetime import datetime


def credit_book_excel(period, cession_in_out_df, cash_pur_df, credit_pur_df,
                      table_sub_total, sub_total, grand_total):
    try:
        # Create a workbook and select the active sheet
        wb = Workbook()
        ws = wb.active

        # Define styles
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add data to specified cells and apply styles
        ws['F1'] = 'SOCAT LLC'
        ws['F1'].font = bold_font
        ws['F1'].alignment = center_alignment

        ws['F2'] = 'OMAN'
        ws['F2'].font = bold_font
        ws['F2'].alignment = center_alignment

        ws['E3'] = 'CREDIT BOOK'
        ws.merge_cells('E3:G3')  # Merge cells E3 to G3 for "CREDIT BOOK"
        ws['E3'].font = bold_font
        ws['E3'].alignment = center_alignment

        ws['E4'] = f'For the Period of {period}'
        ws.merge_cells('E4:G4')  # Merge cells E4 to G4 for the date range
        ws['E4'].font = bold_font
        ws['E4'].alignment = center_alignment

        # Adjust the row height for better appearance
        for row in range(1, 5):
            ws.row_dimensions[row].height = 20

        # Set column width to fit the longest text "Sub Total Cession Out:"
        longest_text = "Sub Total Cession Out:"
        column_width = len(longest_text) + 2  # Add some padding
        for col in range(2, 16):
            ws.column_dimensions[get_column_letter(col)].width = column_width

        # Add headers to the eighth row
        headers = [
            'S.no', 'Narration', 'Food', 'Cleaning', 'Disposal', 'Others',
            'Food', 'Cleaning', 'Disposal', 'Others', 'Food', 'Cleaning',
            'Disposal', 'Others'
        ]

        # Write "Credit Purchase" and DataFrames to the sheet
        start_row = 7
        dfs = [cession_in_out_df, cash_pur_df, credit_pur_df]
        for index, (df, table_total, sub_total_) in enumerate(zip(dfs, table_sub_total, sub_total)):
            # Merge cells and add titles
            ws.merge_cells(start_row=start_row, start_column=3, end_row=start_row, end_column=6)
            ws.cell(row=start_row, column=3, value='Credit Purchase').font = bold_font
            ws.cell(row=start_row, column=3, value='Credit Purchase').alignment = center_alignment
            ws.cell(row=start_row, column=3).border = thin_border

            ws.merge_cells(start_row=start_row, start_column=7, end_row=start_row, end_column=11)
            ws.cell(row=start_row, column=7, value='Cession In').font = bold_font
            ws.cell(row=start_row, column=7, value='Cession In').alignment = center_alignment
            ws.cell(row=start_row, column=7).border = thin_border

            ws.merge_cells(start_row=start_row, start_column=12, end_row=start_row, end_column=15)
            ws.cell(row=start_row, column=12, value='Cession Out').font = bold_font
            ws.cell(row=start_row, column=12, value='Cession Out').alignment = center_alignment
            ws.cell(row=start_row, column=12).border = thin_border

            # Write headers
            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=start_row + 1, column=col_num, value=header)
                cell.font = bold_font
                cell.alignment = center_alignment
                cell.border = thin_border

            # Write DataFrame data
            for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=start_row + 2):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = thin_border

            # Add stock values below the table
            for i in range(12):
                sub_total_value_cell = f'{get_column_letter(i + 3)}{start_row + len(df) + 2}'
                ws[sub_total_value_cell] = table_total[i]
                ws[sub_total_value_cell].font = bold_font
                ws[sub_total_value_cell].alignment = center_alignment
                ws[sub_total_value_cell].border = thin_border

            ws[f'B{start_row + len(df) + 3}'] = 'Sub Total Purchase:'
            ws[f'B{start_row + len(df) + 3}'].font = bold_font
            ws[f'B{start_row + len(df) + 3}'].alignment = center_alignment
            ws[f'B{start_row + len(df) + 3}'].border = thin_border

            ws[f'D{start_row + len(df) + 3}'] = sub_total_[0]
            ws[f'D{start_row + len(df) + 3}'].font = bold_font
            ws[f'D{start_row + len(df) + 3}'].alignment = center_alignment
            ws[f'D{start_row + len(df) + 3}'].border = thin_border

            ws[f'F{start_row + len(df) + 3}'] = 'Sub Total Cession In:'
            ws[f'F{start_row + len(df) + 3}'].font = bold_font
            ws[f'F{start_row + len(df) + 3}'].alignment = center_alignment
            ws[f'F{start_row + len(df) + 3}'].border = thin_border

            ws[f'H{start_row + len(df) + 3}'] = sub_total_[1]
            ws[f'H{start_row + len(df) + 3}'].font = bold_font
            ws[f'H{start_row + len(df) + 3}'].alignment = center_alignment
            ws[f'H{start_row + len(df) + 3}'].border = thin_border

            ws[f'J{start_row + len(df) + 3}'] = 'Sub Total Cession Out:'
            ws[f'J{start_row + len(df) + 3}'].font = bold_font
            ws[f'J{start_row + len(df) + 3}'].alignment = center_alignment
            ws[f'J{start_row + len(df) + 3}'].border = thin_border

            ws[f'L{start_row + len(df) + 3}'] = sub_total_[2]
            ws[f'L{start_row + len(df) + 3}'].font = bold_font
            ws[f'L{start_row + len(df) + 3}'].alignment = center_alignment
            ws[f'L{start_row + len(df) + 3}'].border = thin_border

            start_row += len(
                df) + 4  # Update start_row for next DataFrame, including 1 for "Credit Purchase", 1 for totals, and 2 for spacing

        # Add Grand Totals
        ws[f'B{start_row + 2}'] = 'Grand Total Purchase:'
        ws[f'B{start_row + 2}'].font = bold_font
        ws[f'B{start_row + 2}'].alignment = center_alignment
        ws[f'B{start_row + 2}'].border = thin_border

        ws[f'D{start_row + 2}'] = grand_total[0]
        ws[f'D{start_row + 2}'].font = bold_font
        ws[f'D{start_row + 2}'].alignment = center_alignment
        ws[f'D{start_row + 2}'].border = thin_border

        ws[f'F{start_row + 2}'] = 'Grand Total Cession In:'
        ws[f'F{start_row + 2}'].font = bold_font
        ws[f'F{start_row + 2}'].alignment = center_alignment
        ws[f'F{start_row + 2}'].border = thin_border

        ws[f'H{start_row + 2}'] = grand_total[1]
        ws[f'H{start_row + 2}'].font = bold_font
        ws[f'H{start_row + 2}'].alignment = center_alignment
        ws[f'H{start_row + 2}'].border = thin_border

        ws[f'J{start_row + 2}'] = 'Grand Total Cession Out:'
        ws[f'J{start_row + 2}'].font = bold_font
        ws[f'J{start_row + 2}'].alignment = center_alignment
        ws[f'J{start_row + 2}'].border = thin_border

        ws[f'L{start_row + 2}'] = grand_total[2]
        ws[f'L{start_row + 2}'].font = bold_font
        ws[f'L{start_row + 2}'].alignment = center_alignment
        ws[f'L{start_row + 2}'].border = thin_border

        # Get the current datetime
        current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")

        filename = f"credit_book_{current_datetime}.xlsx"

        # Save the workbook with the current datetime in the filename
        file_path = f'C:/Users/Administrator/Downloads/eiis/credit_book_excel/{filename}'
        wb.save(file_path)
        status = "success"
        print(f"Excel file created and saved at {file_path}")

    except Exception as error:
        print('The cause of error _.', error)
        status = "failed"
        filename = None
        file_path = None

    return status, file_path, filename
