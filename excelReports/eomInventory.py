from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
from datetime import datetime


def create_eom_inv_excel(new_dfs_list, family_name_list, formatted_date, stock_value_total_list,
                         pur_value_total_list, del_value_total_list, cwh_value_total_list,
                         sav_value_total_list, sav_per_value_total_list, total_value_total_list):
    try:
        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active

        # Set values in the specified cells
        ws['H1'] = "SOCAT LLC"
        ws['H2'] = "OMAN"
        ws['F3'] = f"EOM INVENTORY FOR THE PERIOD OF - {formatted_date}"

        # Merge cells from F3 to K3
        ws.merge_cells('F3:K3')

        # Center align the merged cell content and set font to bold
        alignment = Alignment(horizontal='center')
        bold_font = Font(bold=True)

        # Define border style
        border = Border(left=Side(border_style='thin'),
                        right=Side(border_style='thin'),
                        top=Side(border_style='thin'),
                        bottom=Side(border_style='thin'))

        # Apply center alignment and bold font to F3:K3
        for col in range(6, 12):  # F is the 6th column, K is the 11th
            col_letter = get_column_letter(col)
            cell = ws[f'{col_letter}3']
            cell.alignment = alignment
            cell.font = bold_font

        # Apply bold font and center alignment to the other specified cells
        ws['H1'].alignment = alignment
        ws['H1'].font = bold_font

        ws['H2'].alignment = alignment
        ws['H2'].font = bold_font

        # Starting row for the first family name and table
        start_row = 5

        # Iterate through family names and DataFrames
        for family_name, df, stock_value, pur_val, del_val, cwh_val, sav_val, sav_per_val, total_val in zip(
                family_name_list, new_dfs_list, stock_value_total_list, pur_value_total_list,
                del_value_total_list,
                cwh_value_total_list,
                sav_value_total_list,
                sav_per_value_total_list,
                total_value_total_list):
            # Print "Family Name: " followed by the family name
            family_cell = f'A{start_row}'
            stock_cell = f'D{start_row}'
            purchase_cell = f'F{start_row}'
            delivery_cell = f'H{start_row}'
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
            ws[family_cell] = f"Family Name: {family_name}"
            ws[family_cell].font = bold_font
            ws[family_cell].alignment = alignment
            ws[family_cell].border = border  # Apply border to family name cell

            # Add additional texts with bold formatting and borders
            ws.merge_cells(start_row=start_row, start_column=4, end_row=start_row, end_column=5)
            ws[stock_cell] = 'Stock BOM'
            ws[stock_cell].font = bold_font
            ws[stock_cell].alignment = alignment
            ws[stock_cell].border = border  # Apply border to Stock BOM cell

            ws.merge_cells(start_row=start_row, start_column=6, end_row=start_row, end_column=7)
            ws[purchase_cell] = 'Purchase'
            ws[purchase_cell].font = bold_font
            ws[purchase_cell].alignment = alignment
            ws[purchase_cell].border = border  # Apply border to Purchase cell

            ws.merge_cells(start_row=start_row, start_column=8, end_row=start_row, end_column=11)
            ws[delivery_cell] = 'Delivery'
            ws[delivery_cell].font = bold_font
            ws[delivery_cell].alignment = alignment
            ws[delivery_cell].border = border  # Apply border to Delivery cell

            # Print DataFrame column names
            header_start_row = start_row + 1
            for col_idx, col_name in enumerate(df.columns, start=1):
                header_cell = ws.cell(row=header_start_row, column=col_idx, value=col_name)
                header_cell.font = bold_font
                header_cell.alignment = alignment
                header_cell.border = border  # Apply border to header cells

            # Start writing the DataFrame below the column headers
            table_start_row = header_start_row + 1
            for r_idx, row in enumerate(df.itertuples(index=False), start=table_start_row):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = border  # Apply border to each cell

            # Apply border to table headers
            for col_idx in range(1, len(df.columns) + 1):
                for row_idx in [header_start_row]:
                    ws.cell(row=row_idx, column=col_idx).border = border

            # Apply border to table rows
            for col_idx in range(1, len(df.columns) + 1):
                for row_idx in range(header_start_row + 1, table_start_row + len(df)):
                    ws.cell(row=row_idx, column=col_idx).border = border

            # Add stock value below the table in the E column
            sub_total_value_cell = f'C{table_start_row + len(df) + 1}'
            ws[sub_total_value_cell] = 'Sub Total :'
            ws[sub_total_value_cell].font = bold_font
            ws[sub_total_value_cell].alignment = alignment
            ws[sub_total_value_cell].border = border

            stock_value_cell = f'E{table_start_row + len(df) + 1}'
            ws[stock_value_cell] = stock_value
            ws[stock_value_cell].font = bold_font
            ws[stock_value_cell].alignment = alignment
            ws[stock_value_cell].border = border

            purchase_value_cell = f'G{table_start_row + len(df) + 1}'
            ws[purchase_value_cell] = pur_val
            ws[purchase_value_cell].font = bold_font
            ws[purchase_value_cell].alignment = alignment
            ws[purchase_value_cell].border = border

            delivery_value_cell = f'I{table_start_row + len(df) + 1}'
            ws[delivery_value_cell] = del_val
            ws[delivery_value_cell].font = bold_font
            ws[delivery_value_cell].alignment = alignment
            ws[delivery_value_cell].border = border

            cwh_value_cell = f'L{table_start_row + len(df) + 1}'
            ws[cwh_value_cell] = cwh_val
            ws[cwh_value_cell].font = bold_font
            ws[cwh_value_cell].alignment = alignment
            ws[cwh_value_cell].border = border

            saving_value_cell = f'M{table_start_row + len(df) + 1}'
            ws[saving_value_cell] = sav_val
            ws[saving_value_cell].font = bold_font
            ws[saving_value_cell].alignment = alignment
            ws[saving_value_cell].border = border

            saving_percentage_value_cell = f'N{table_start_row + len(df) + 1}'
            ws[saving_percentage_value_cell] = sav_per_val
            ws[saving_percentage_value_cell].font = bold_font
            ws[saving_percentage_value_cell].alignment = alignment
            ws[saving_percentage_value_cell].border = border

            total_value_cell = f'P{table_start_row + len(df) + 1}'
            ws[total_value_cell] = total_val
            ws[total_value_cell].font = bold_font
            ws[total_value_cell].alignment = alignment
            ws[total_value_cell].border = border

            # Adjust column widths based on the maximum length in each column, including headers
            for col_idx, col_name in enumerate(df.columns, start=1):
                max_length = max(len(col_name), max(len(str(value)) for value in df.iloc[:, col_idx - 1]))
                adjusted_width = max_length + 2  # Adding extra space for better readability
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = adjusted_width

            # Update the start_row for the next family name and table
            start_row = table_start_row + len(df) + 3  # +3 for spacing between tables and stock value row

        # Add Grand Total values below all tables
        sub_total_value_cell = f'C{start_row}'
        ws[sub_total_value_cell] = 'Grand Total :'
        ws[sub_total_value_cell].font = bold_font
        ws[sub_total_value_cell].alignment = alignment
        ws[sub_total_value_cell].border = border

        stock_value_total = round(sum(stock_value_total_list), 3)
        purchase_value_total = round(sum(pur_value_total_list), 3)
        delivery_value_total = round(sum(del_value_total_list), 3)
        cwh_value_total = round(sum(cwh_value_total_list), 3)
        saving_value_total = round(sum(sav_value_total_list), 3)
        saving_percentage_value_total = round(sum(sav_per_value_total_list), 3)
        total_value_total = round(sum(total_value_total_list), 3)

        # Update column widths based on content
        column_widths = {
            'E': len(str(stock_value_total)) + 2,
            'G': len(str(purchase_value_total)) + 2,
            'I': len(str(delivery_value_total)) + 2,
            'L': len(str(cwh_value_total)) + 2,
            'M': len(str(saving_value_total)) + 2,
            'N': len(str(saving_percentage_value_total)) + 2,
            'P': len(str(total_value_total)) + 2,
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        ws[f'E{start_row}'] = stock_value_total
        ws[f'G{start_row}'] = purchase_value_total
        ws[f'I{start_row}'] = delivery_value_total
        ws[f'L{start_row}'] = cwh_value_total
        ws[f'M{start_row}'] = saving_value_total
        ws[f'N{start_row}'] = saving_percentage_value_total
        ws[f'P{start_row}'] = total_value_total

        # Set styles for Grand Total values
        for col in ['E', 'G', 'I', 'L', 'M', 'N', 'P']:
            cell = ws[f'{col}{start_row}']
            cell.font = bold_font
            cell.alignment = alignment
            cell.border = border

        # Generate a filename with the current timestamp
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"EOM_INV_{current_time}.xlsx"

        # Save the workbook in the specified directory
        file_path = f"C:/Users/Administrator/Downloads/eiis/EOM_INV_EXCEL/{filename}"
        wb.save(file_path)

        print(f"File saved as: {file_path}")
        status = 'success'
    except Exception as error:
        print('The cause of error ->', error)
        status = 'failed'
        filename = None
        file_path = None

    return status, file_path, filename
