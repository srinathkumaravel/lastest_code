import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime


def create_cwh_savings_excel(df, formatted_date):
    try:
        # Create a workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Define the data and cell positions
        header_data = [
            ("C1", "SOCAT LLC"),
            ("C2", "OMAN"),
            ("C3", "CWH SAVINGS"),
            ("C4", f"FOR THE PERIOD {formatted_date}"),
        ]

        # Define the style for bold font and central alignment
        bold_centered = Font(bold=True)
        center_alignment = Alignment(horizontal='center')

        # Write the header data to the worksheet and apply the styles
        for cell, text in header_data:
            ws[cell] = text
            ws[cell].font = bold_centered
            ws[cell].alignment = center_alignment

        # Define the style for bold font for the table headers
        bold_font = Font(bold=True)

        # Define border styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write the column headers with bold font and borders
        for col_idx, col_name in enumerate(df.columns, start=2):
            cell = ws.cell(row=5, column=col_idx, value=col_name)
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # Write DataFrame to the worksheet starting from the 6th row
        for row_idx, (index, row) in enumerate(df.iterrows(), start=6):
            for col_idx, value in enumerate(row, start=2):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Add the "Total" label and calculate the totals
        total_row = df.shape[0] + 6
        ws[f"B{total_row}"] = "Total"
        ws[f"B{total_row}"].font = bold_font

        # Calculate the total for the "Savings" column and other numeric columns
        for col_idx, col_name in enumerate(df.columns[2:], start=4):  # start=4 to skip to numeric columns
            total = df[col_name].sum()
            ws.cell(row=total_row, column=col_idx, value=total).font = bold_font

        # Adjust column widths based on the length of the text in each cell
        for col in ws.columns:
            max_length = 0
            column = [cell for cell in col if cell.value is not None]  # Filter out cells with no value
            if column:  # Check if column is not empty
                for cell in column:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))  # Convert cell value to string
                adjusted_width = max_length + 2  # Adding a little extra space
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Get the current time and format it
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = "CWH_SAVINGS_{current_time}.xlsx"
        # Define the directory and ensure it exists
        file_path = fr'C:/Users/Administrator/Downloads/eiis/CWH_SAVINGS_EXCEL/{file_name}'

        # Save the workbook
        wb.save(file_path)
        status = "success"
        print(f"Excel file created and saved as {file_path}")
    except Exception as error:
        print('The cause of error -->', error)
        status = "failed"
        file_name = None
        file_path = None

    return status, file_name, file_path
