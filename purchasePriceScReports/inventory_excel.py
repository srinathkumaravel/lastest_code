import pandas as pd
from database import get_database_engine_e_eiis
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import MergedCell  # Import MergedCell


def fetch_inventory_data(from_date, to_date, excel_type, file_path):
    try:
        engine = get_database_engine_e_eiis()
        inventory_query = """
                            SELECT 
                                PERIOD, 
                                CATEGORY_NAME, 
                                TOTAL_QTY 
                            FROM 
                                inventory 
                            WHERE 
                                PERIOD BETWEEN %s AND %s 
                                AND REPORT_TYPE = %s;
                            """
        # Replace database_connection with your actual database connection object
        df = pd.read_sql_query(inventory_query, engine, params=(from_date, to_date, excel_type))
        # Convert PERIOD column to datetime format
        df['PERIOD'] = pd.to_datetime(df['PERIOD'])
        # Ensure all months within the range are included
        date_range = pd.date_range(start=from_date, end=to_date, freq='ME').strftime('%b-%y')
        for month_year in date_range:
            if month_year not in df['PERIOD'].dt.strftime('%b-%y').unique():
                # Create a new row with zeros for the missing month
                zero_row = pd.DataFrame({
                    'PERIOD': pd.to_datetime([f'01-{month_year}']),
                    'CATEGORY_NAME': 'N/A',
                    'TOTAL_QTY': 0
                })
                df = pd.concat([df, zero_row], ignore_index=True)

        # Extract month and year as strings like 'jun-24', 'may-24', etc.
        df['Month_Year'] = df['PERIOD'].dt.strftime('%b-%y')

        # Pivot the data to create separate columns for QTY, Amt, and AvgUnitPrice for each Month_Year
        pivot_df = df.pivot_table(
            index=['CATEGORY_NAME'],
            columns='Month_Year',
            values=['TOTAL_QTY'],
            aggfunc='sum').reset_index()
        # Flatten the column names to remove the multi-index
        pivot_df.columns = ['_'.join(col).strip() for col in pivot_df.columns.values]

        # Replace NaN values with 0
        pivot_df = pivot_df.fillna(0)
        pivot_df = pivot_df[pivot_df['CATEGORY_NAME_'] != 'N/A']
        # Remove 'TOTAL_QTY_' from column names if it exists
        pivot_df.columns = [col.replace('TOTAL_QTY_', '') for col in pivot_df.columns]
        pivot_df.columns = [col.replace('CATEGORY_NAME_', 'Category Name') for col in pivot_df.columns]
        # Filter out the 'Category Name' column and select numeric columns
        numeric_columns = pivot_df.select_dtypes(include='number').columns

        # Calculate the total for each row across the selected numeric columns
        pivot_df.loc[:, 'Total'] = pivot_df[numeric_columns].sum(axis=1).round(3)

        # Calculate the average for each row across the selected numeric columns
        pivot_df.loc[:, 'Average'] = pivot_df[numeric_columns].mean(axis=1).round(3)

        status, message = create_inventory_excel(pivot_df, file_path)

    except Exception as error:
        print('The cause of error -->', error)
        status = 'failed'
        message = 'failed'

    return status, message


def create_inventory_excel(pivot_df, file_path):
    try:
        # Load the existing workbook
        workbook = load_workbook(file_path)

        # Add a new sheet named 'inventory'
        if 'inventory' in workbook.sheetnames:
            sheet = workbook['inventory']
        else:
            sheet = workbook.create_sheet(title='inventory')

        # Add title in bold font across A1 to E1
        title = 'CATEGORY WISE INVENTORY'
        sheet.merge_cells('A1:E1')  # Merge cells A1 to E1
        cell = sheet['A1']
        cell.value = title
        cell.font = Font(bold=True)  # Set font to bold
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Center the text

        # Add the pivot_df starting from row 4, column A
        for r_idx, row in enumerate(dataframe_to_rows(pivot_df, index=False, header=True), start=4):
            for c_idx, value in enumerate(row, start=1):  # Start from column A (1)
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)

                # Apply borders
                border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                cell.border = border

        # Apply bold font and center alignment to header row
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')

        for cell in sheet[4]:  # Header row is row 4
            cell.font = header_font
            cell.alignment = center_alignment

        # Apply background colors
        light_grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light Grey
        light_red_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")  # Light Red 4
        light_green_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")  # Light Green 4

        # Determine the columns to apply each color
        total_columns = len(pivot_df.columns)
        last_two_cols = total_columns - 2

        # Apply Light Grey to the first column
        for row in sheet.iter_rows(min_row=4, max_col=1, max_row=sheet.max_row):
            for cell in row:
                cell.fill = light_grey_fill

        # Apply Light Red to all columns except the last two
        for col_idx in range(2, last_two_cols + 1):  # Columns from B to last_two_cols
            for row in sheet.iter_rows(min_row=4, max_col=col_idx, max_row=sheet.max_row):
                for cell in row:
                    cell.fill = light_red_fill

        # Apply Light Green to the last two columns
        for col_idx in range(last_two_cols + 1, total_columns + 1):  # Last two columns
            for row in sheet.iter_rows(min_row=4, max_col=col_idx, max_row=sheet.max_row):
                for cell in row:
                    cell.fill = light_green_fill

        # Adjust column widths based on the content
        for col in sheet.columns:
            max_length = 0
            column_letter = None
            for cell in col:
                if not isinstance(cell, MergedCell):  # Skip merged cells
                    if column_letter is None:
                        column_letter = cell.column_letter  # Get the column letter from the first non-merged cell
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            if column_letter:
                adjusted_width = (max_length + 2)  # Add some extra space
                sheet.column_dimensions[column_letter].width = adjusted_width

        # Calculate and add totals for columns except the first and last two columns
        # Find the row where the table ends
        end_row = sheet.max_row + 2

        # Exclude the first and last two columns from total calculation
        columns_to_sum = pivot_df.columns[1:]

        # Write totals
        sheet.cell(row=end_row, column=1, value='Total').font = Font(bold=True)  # Label for totals
        for idx, col in enumerate(columns_to_sum, start=2):
            column_letter = sheet.cell(row=4, column=idx).column_letter  # Get the column letter
            total_formula = f'=SUM({column_letter}4:{column_letter}{end_row - 2})'  # Create the SUM formula
            cell = sheet.cell(row=end_row, column=idx, value=total_formula)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin')
            )

        # Save the workbook with the new sheet and data
        workbook.save(file_path)
        status = 'success'
        message = 'success'
    except Exception as error:
        print('The cause of error -->', error)
        status = 'failed'
        message = 'success'

    return status, message
