import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl  # Ensure openpyxl is imported
from database import get_database_engine_e_eiis


def fetch_theoritical_data():
    dfs_list = []
    family_id_list = []
    family_name_list = []
    total_qty_list = []
    total_value_list = []
    status = "failed"
    try:
        engine = get_database_engine_e_eiis()
        # Query to fetch item categories
        df_query = """SELECT ITEM_CAT_PK, NAME FROM mst_item_category;"""
        dict_df = pd.read_sql_query(df_query, engine)

        # Initialize an empty dictionary
        family_names = {}

        # Populate dictionary with item categories
        for index, row in dict_df.iterrows():
            key = str(row['ITEM_CAT_PK'])
            value = row['NAME']
            family_names[key] = value

        # Query to fetch item data
        data_query = """SELECT st.ITEM_ID, item.Item_Name, st.PACKAGE_ID, st.IN_QTY, 
        st.IN_GP FROM stock AS st INNER JOIN item ON item.Item_ID = st.ITEM_ID;"""
        data_df = pd.read_sql_query(data_query, engine)

        # Add 'Family ID' column based on the first two characters of 'ITEM_ID'
        data_df['Family ID'] = data_df['ITEM_ID'].astype(str).str[:2]
        # Convert column 'A' from int to str
        data_df['ITEM_ID'] = data_df['ITEM_ID'].astype(str)

        # Map 'Family ID' to 'Family Name' using the family_names dictionary
        data_df['Family Name'] = data_df['Family ID'].map(family_names)

        # Splitting data_df based on 'Family ID'
        grouped = data_df.groupby('Family ID')

        # Initializing an empty list to store each DataFrame
        dfs_list = []
        family_id_list = []
        family_name_list = []
        total_qty_list = []
        total_value_list = []

        # Iterate through each group, creating a DataFrame for each unique 'Family ID'
        for group_name, group_df in grouped:
            # Assuming 'df' is your DataFrame
            # Accessing the first row, sixth column value
            family_ID = group_df.iloc[0, 5]  # Note: index 5 because it's 0-based
            family_Name = group_df.iloc[0, 6]
            total_qty_list.append(round(group_df['IN_QTY'].sum(), 2))
            total_value_list.append(round(group_df['IN_GP'].sum(), 2))
            family_id_list.append(family_ID)
            family_name_list.append(family_Name)
            group_df.drop(columns=['Family Name', 'Family ID'], inplace=True)
            # Define a dictionary with old names as keys and new names as values
            rename_dict = {
                'ITEM_ID': 'Item Code',
                'Item_Name': 'Item Name',
                'PACKAGE_ID': 'Packing',
                'IN_QTY': 'Quantity',
                'IN_GP': 'Value'
            }

            # Rename the columns using the dictionary
            group_df.rename(columns=rename_dict, inplace=True)
            dfs_list.append(group_df)
            status = "success"
    except Exception as error:
        print('The cause of error -->', error)
        status = "failed"

    return dfs_list, family_id_list, family_name_list, total_qty_list, total_value_list, status


def generate_theoritical_excel_report(dfs_list, family_id_list, family_name_list, total_qty_list, total_value_list):
    excel_filepath = None
    file_name = None
    try:
        # Create a workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Load your image
        image_path = 'C:\\Users\\Administrator\\Downloads\\eiis\\sodexo.jpg'
        img = XLImage(image_path)
        img.width = 200  # Adjust width to cover multiple cells
        img.height = 100  # Adjust height to cover multiple cells

        # Add image to start cell (A1)
        ws.add_image(img, 'A1')

        # Merge cells A1 to D2 to cover the image
        ws.merge_cells('A1:D2')

        # Define text for cells E3 and E4
        text1 = "SOCAT LLC - OMAN"
        text2 = "Theoretical Stock"

        # Assign text to cell E3
        ws['E3'] = text1

        # Assign text to cell E4
        ws['E4'] = text2

        # Merge cells E3 to J3
        ws.merge_cells('E3:J3')

        # Center align the text in the merged cell E3:J3
        merged_cell_E3 = ws['E3']
        merged_cell_E3.alignment = Alignment(horizontal='center', vertical='center')

        # Apply bold and increase font size for text in E3:J3
        font_E3 = Font(bold=True, size=14)
        for row in ws['E3:J3']:
            for cell in row:
                cell.font = font_E3

        # Merge cells E4 to J4
        ws.merge_cells('E4:J4')

        # Center align the text in the merged cell E4:J4
        merged_cell_E4 = ws['E4']
        merged_cell_E4.alignment = Alignment(horizontal='center', vertical='center')

        # Apply bold and increase font size for text in E4:J4
        font_E4 = Font(bold=True, size=14)
        for row in ws['E4:J4']:
            for cell in row:
                cell.font = font_E4

        # Add thick black borderlines around the rectangle from A1 to J7
        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )

        # Apply border to the entire perimeter from A1 to J7
        for row in range(1, 8):  # Rows 1 to 7
            for col in range(1, 11):  # Columns A to J
                cell = ws.cell(row=row, column=col)
                if row == 1:  # Top row
                    cell.border = Border(top=border.top)
                if row == 7:  # Bottom row
                    cell.border = Border(bottom=border.bottom)
                if col == 1:  # Left column
                    cell.border = Border(left=border.left)
                if col == 10:  # Right column
                    cell.border = Border(right=border.right)

        # Set specific widths for columns A to E
        column_widths = {'A': 15, 'B': 21, 'C': 15, 'D': 15, 'E': 20}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Starting row for the next table (leave an empty row)
        start_row = 9
        Grand_total = 0
        # Iterate through each set of values and corresponding DataFrame using zip
        for df, family_id, family_name, total_qty, total_value in zip(dfs_list, family_id_list, family_name_list,
                                                                      total_qty_list, total_value_list):
            # Insert empty row if it's not the first table
            Grand_total += total_value
            if start_row > 9:
                start_row += 2  # Move down by 2 rows to leave empty space

            # Write Family ID and ID
            ws.cell(row=start_row, column=1, value='Family ID :')
            ws.cell(row=start_row, column=1).font = Font(bold=True)  # Set the font to bold

            ws.cell(row=start_row, column=2, value=family_id)

            # Leave two empty cells
            ws.cell(row=start_row, column=3)
            # ws.cell(row=start_row, column=4)

            # Write Family Name and Value
            ws.cell(row=start_row, column=4, value='Family Name :')
            ws.cell(row=start_row, column=4).font = Font(bold=True)
            ws.cell(row=start_row, column=5, value=family_name)

            # Add the DataFrame table starting from column A
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row + 2):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == start_row + 2:
                        cell.font = Font(bold=True)  # Bold font for header row

                    # Apply borders to all cells in the table
                    cell.border = border

            # Add Total Qty and Total Value after the table
            total_cell = ws.cell(row=start_row + len(df) + 3, column=3, value='Total :')
            total_cell.font = Font(bold=True)

            ws.cell(row=start_row + len(df) + 3, column=4, value=total_qty)
            ws.cell(row=start_row + len(df) + 3, column=5, value=total_value)

            # Increment start_row for the next table
            start_row += len(df) + 3  # Move down by the length of the dataframe + 3 rows for spacing

        ws.cell(row=start_row + 1, column=4, value='Grand Total :')
        ws.cell(row=start_row + 1, column=4).font = Font(bold=True)  # Set the font to bold
        ws.cell(row=start_row + 1, column=5, value=Grand_total)
        print(Grand_total)
        # Get current time to add to the filename
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        file_name = f"Theoritical_stock_{current_time}.xlsx"
        # Save the workbook with current time added to the filename
        excel_filepath = rf'C:\Users\Administrator\Downloads\eiis\theoreticalsupplier\{file_name}'
        wb.save(excel_filepath)
        status = "success"
    except Exception as error:
        print('The cause of error -->', error)
        status = 'failed'

    return status, file_name, excel_filepath
