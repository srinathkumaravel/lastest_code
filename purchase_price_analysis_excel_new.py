import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from database import get_database_engine_e_eiis
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def get_data_for_excel(from_date, to_date, excel_type):
    def convert_to_month_end(date_str):
        # Convert the date string to a datetime object
        date = pd.to_datetime(date_str)
        # Use MonthEnd to get the last day of the month
        month_end = date + pd.offsets.MonthEnd(0)
        return month_end

    to_date = convert_to_month_end(to_date)

    concatenated_df = None
    desired_order: list = []
    try:
        # purchase_price_analysis_cash_purchase
        if excel_type == "1":
            engine = get_database_engine_e_eiis()
            # Query to fetch item categories
            cash_purchase_query = """SELECT 
                                            PERIOD, 
                                                    category_name, 
                                                    category_id, 
                                                    ITEM_ID, 
                                                    ITEM_NAME, 
                                                    PACKAGE_ID, 
                                                    UOM,
                                                    QTY, 
                                                    Amt, 
                                                    AvgUnitPrice 
                                                FROM 
                                                    purchase_price_analysis_cash_purchase 
                                                WHERE 
                                                    PERIOD BETWEEN %s AND %s;
                                                """
            # Replace database_connection with your actual database connection object
            df = pd.read_sql_query(cash_purchase_query, engine, params=(from_date, to_date))

        # purchase_price_analysis_out_of_catalogue_item

        elif excel_type == "2":
            engine = get_database_engine_e_eiis()
            out_of_catalogue_item_query = """SELECT 
                                                    PERIOD, 
                                                    category_name, 
                                                    category_id, 
                                                    ITEM_CODE, 
                                                    ITEM_NAME, 
                                                    PACKAGE_ID, 
                                                    UOM,
                                                    QTY, 
                                                    Amt, 
                                                    AvgUnitPrice 
                                                FROM 
                                                    purchase_price_analysis_out_of_catalogue_item 
                                                WHERE 
                                                    PERIOD BETWEEN %s AND %s;
                                                """
            # Replace database_connection with your actual database connection object
            df = pd.read_sql_query(out_of_catalogue_item_query, engine, params=(from_date, to_date))
            # Rename the column
            df.rename(columns={'ITEM_CODE': 'ITEM_ID'}, inplace=True)

        else:
            engine = get_database_engine_e_eiis()
            # Query to fetch purchase price analysis
            purchase_price_analysis_query = """SELECT 
                                                    PERIOD, 
                                                    category_name, 
                                                    category_id, 
                                                    ITEM_ID, 
                                                    ITEM_NAME, 
                                                    PACKAGE_ID, 
                                                    UOM,
                                                    QTY, 
                                                    Amt, 
                                                    AvgUnitPrice 
                                                FROM 
                                                    purchase_price_analysis 
                                                WHERE 
                                                    PERIOD BETWEEN %s AND %s;
                                                """
            # Replace database_connection with your actual database connection object
            df = pd.read_sql_query(purchase_price_analysis_query, engine, params=(from_date, to_date))
        if len(df) != 0:
            # Convert PERIOD column to datetime format
            df['PERIOD'] = pd.to_datetime(df['PERIOD'])

            # Ensure all months within the range are included
            date_range = pd.date_range(start=from_date, end=to_date, freq='ME').strftime('%b-%y')
            for month_year in date_range:
                if month_year not in df['PERIOD'].dt.strftime('%b-%y').unique():
                    # Create a new row with zeros for the missing month
                    zero_row = pd.DataFrame({
                        'PERIOD': pd.to_datetime([f'01-{month_year}']),
                        'category_name': 'N/A',
                        'category_id': 'N/A',
                        'ITEM_ID': 'N/A',
                        'ITEM_NAME': 'N/A',
                        'PACKAGE_ID': 'N/A',
                        'UOM': 'N/A',
                        'QTY': 0,
                        'Amt': 0,
                        'AvgUnitPrice': 0
                    })
                    df = pd.concat([df, zero_row], ignore_index=True)

            # Extract month and year as strings like 'jun-24', 'may-24', etc.
            df['Month_Year'] = df['PERIOD'].dt.strftime('%b-%y')

            # Pivot the data to create separate columns for QTY, Amt, and AvgUnitPrice for each Month_Year
            pivot_df = df.pivot_table(
                index=['category_name', 'category_id', 'ITEM_ID', 'ITEM_NAME', 'PACKAGE_ID', 'UOM'],
                columns='Month_Year',
                values=['QTY', 'Amt', 'AvgUnitPrice'],
                aggfunc='sum').reset_index()

            # Flatten the column names to remove the multi-index
            pivot_df.columns = ['_'.join(col).strip() for col in pivot_df.columns.values]

            # Replace NaN values with 0
            pivot_df = pivot_df.fillna(0)

            # Define a regex pattern to match month-year identifiers (e.g., Feb-24, Mar-24, etc.)
            pattern = r'([A-Za-z]{3}-\d{2})'  # Matches month abbreviation followed by two digits

            # Extract all column names from the pivot_df DataFrame
            all_columns = pivot_df.columns.tolist()

            # Dictionary to store groups based on month-year identifiers
            groups = {}

            # Iterate through all columns and categorize based on the regex pattern
            for column in all_columns:
                match = re.search(pattern, column)
                if match:
                    month_year = match.group(1)  # Extract the matched month-year identifier
                    if month_year not in groups:
                        groups[month_year] = []
                    groups[month_year].append(column)

            # Create a date range with monthly frequency
            date_range = pd.date_range(start=from_date, end=to_date, freq='M')

            # Format the dates to the desired 'Mon-YY' format
            desired_order = date_range.strftime('%b-%y').tolist()
            # Reverse the list
            desired_order = desired_order[::-1]
            print(desired_order)
            sorted_columns = []
            for month in desired_order:
                if month in groups:
                    sorted_columns.extend(groups[month])

            # Define the set of initial columns
            initial_columns = ['category_name_', 'category_id_', 'ITEM_ID_', 'ITEM_NAME_', 'PACKAGE_ID_', 'UOM_']
            # Ensure initial columns are added at the beginning of the sorted_columns list
            sorted_columns = initial_columns + sorted_columns

            # Reorder columns in the pivot_df DataFrame
            pivot_df = pivot_df[sorted_columns]

            amt_columns = [col for col in pivot_df.columns if col.startswith('Amt_')]
            pivot_df['Amt_mean'] = pivot_df[amt_columns].mean(axis=1).round(3)
            avg_unit_price_columns = [col for col in pivot_df.columns if col.startswith('AvgUnitPrice_')]
            pivot_df['AvgUnitPrice_mean'] = pivot_df[avg_unit_price_columns].mean(axis=1).round(3)
            qty_columns = [col for col in pivot_df.columns if col.startswith('QTY_')]
            pivot_df['QTY_mean'] = pivot_df[qty_columns].mean(axis=1).round(3)

            # Get column names from 6th column onward
            column_names = pivot_df.columns[6:].tolist()
            print(column_names)
            # Convert columns from 6th onwards to lists
            columns_to_lists = pivot_df.columns[6:]  # Select columns from 6th column onwards
            lists = [pivot_df[col].tolist() for col in columns_to_lists]

            val = 0
            for i, (lst, name) in enumerate(zip(lists, column_names)):
                val += 1  # Increment val (counter)
                if val % 3 == 0:
                    # print(lst)  # Print the list element
                    print('yes')
                    Qty_column = lists.pop(i)  # Remove element from lists and store in Qty_column
                    column_name = column_names.pop(i)  # Remove element from column_names and store in column_name
                    i -= 2  # Decrement i to adjust for the list modifications

                    # Insert removed elements back at adjusted position
                    lists.insert(i, Qty_column)
                    column_names.insert(i, column_name)

            # Extract first 6 columns
            first_six_columns = pivot_df.iloc[:, :6]

            # Determine the maximum length of any sublist (assuming all columns have the same length)
            max_length = max(len(lst) for lst in lists)

            # Pad shorter lists with NaNs (if needed to ensure equal length)
            for i, lst in enumerate(lists):
                if len(lst) < max_length:
                    lists[i].extend([float('nan')] * (max_length - len(lst)))

            # Create DataFrame
            month_df = pd.DataFrame(lists).transpose()
            month_df.columns = column_names

            # Concatenate along columns (horizontally)
            concatenated_df = pd.concat([first_six_columns, month_df], axis=1)

            # Function to rename columns based on prefixes
            def rename_columns(col):
                if col.startswith('QTY_'):
                    return 'Qty'  # Append the rest of the column name after 'QTY_'
                elif col.startswith('Amt_'):
                    return 'Amount'  # Append the rest of the column name after 'Amt_'
                elif col.startswith('Avg'):
                    return 'Avg Unit Price'  # Append the rest of the column name after 'Avg_'
                else:
                    return col  # Return the original column name if no prefix match

            # Apply renaming function to all columns
            concatenated_df.rename(columns=rename_columns, inplace=True)
            # Define renaming rules in a single dictionary
            rename_dict = {
                'category_name_': 'Category Name',
                'category_id_': 'Category ID',
                'ITEM_ID_': 'Item Code',
                'ITEM_NAME_': 'Item Name',
                'PACKAGE_ID_': 'Package ID',
                'UOM_': 'UOM'}

            # Apply renaming to the DataFrame
            concatenated_df.rename(columns=rename_dict, inplace=True)
            status = "success"
            message = "success"
        else:
            print(f'No data available ofr the selected period from {from_date} to {to_date}')
            status = "success"
            message = "No data available"
            return concatenated_df, desired_order, message, status

    except Exception as error:
        print('The cause of error -->', error)
        status = "failed"
        message = "failed"

    return concatenated_df, desired_order, message, status


def create_purchase_price_excel_report(concatenated_df, desired_order, excel_type):
    file_name = None
    file_path = None

    try:
        # Get the current date and time
        current_datetime = datetime.now().strftime('%Y%m%d_%H%M%S')
        if excel_type == "1":
            file_name = f"PurchaseAnalysisReport_CashPurchaseItems_{current_datetime}.xlsx"
        elif excel_type == "2":
            file_name = f"PurchaseAnalysisReport_OutOfCatalogueItems_{current_datetime}.xlsx"
        else:
            file_name = f"PurchaseAnalysisReport_{current_datetime}.xlsx"
        # Create a filename with the current date and time
        file_path = fr"C:\Users\Administrator\Downloads\eiis\purchase_analysis_excel\{file_name}"

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Create an initial blank sheet
            pd.DataFrame().to_excel(writer, index=False, header=False)

        # Load the workbook and select the active sheet
        wb = load_workbook(file_path)
        ws = wb.active

        # Write the DataFrame to the 6th row
        for r_idx, row in enumerate(dataframe_to_rows(concatenated_df, index=False, header=True), 6):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Merge cells and add values from the desired_order list
        start_column = 7  # G column
        start_row = 5
        columns_per_month = 3  # Number of columns to merge for each month
        length_of_desired_order = len(desired_order)
        desired_order.append(f'Avg of {length_of_desired_order} month')
        for idx, month in enumerate(desired_order):
            col_start = start_column + (idx * columns_per_month)
            col_end = col_start + columns_per_month - 1
            merge_range = f"{get_column_letter(col_start)}{start_row}:{get_column_letter(col_end)}{start_row}"
            ws.merge_cells(merge_range)
            ws[f"{get_column_letter(col_start)}{start_row}"].value = month
            ws[f"{get_column_letter(col_start)}{start_row}"].alignment = Alignment(horizontal='center',
                                                                                   vertical='center')

        # Freeze the first 6 rows and columns A to F
        ws.freeze_panes = 'G7'  # This will freeze rows 1 to 6 and columns A to F

        # Add sky blue color to cells in columns A to F for rows 5 and 6
        fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
        for row in [5, 6]:
            for col in range(1, 7):  # Columns A to F
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.font = Font(bold=True, color='FFFFFF')  # Bold font and white color

        # Define colors for each group of 3 columns
        colors = ['C6EFCE', 'FCE4D6', 'D9D9D9']  # Add more colors as needed

        # Apply color to columns based on desired_order
        for idx, month in enumerate(desired_order):
            color_index = idx % len(colors)  # Cycle through colors
            col_start = start_column + (idx * columns_per_month)
            col_end = col_start + columns_per_month - 1
            color = colors[color_index]

            for col_idx in range(col_start, col_end + 1):
                for row_idx in range(start_row, ws.max_row + 1):
                    col_letter = get_column_letter(col_idx)
                    cell = ws[f"{col_letter}{row_idx}"]
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        # Define border style
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Apply border to the range of cells containing data
        data_start_row = 6  # Starting row where data begins (after headers)
        data_end_row = ws.max_row  # Last row with data
        data_start_col = 1  # Starting column for data (column A)
        data_end_col = ws.max_column  # Last column with data

        for row_idx in range(data_start_row, data_end_row + 1):
            for col_idx in range(data_start_col, data_end_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border

        # Add 'SOCAT' to the first row, D column, with bold font, font size 10, and center alignment
        socat_cell = ws.cell(row=1, column=4, value='SOCAT')
        socat_cell.font = Font(bold=True, size=10)
        socat_cell.alignment = Alignment(horizontal='center', vertical='center')

        if excel_type == "2":
            socat_cell = ws.cell(row=2, column=4, value='Purchase Analysis Report (Out of Catalogue Items)')
        elif excel_type == "1":
            socat_cell = ws.cell(row=2, column=4, value='Purchase Analysis Report (Cash Purchase Items)')
        else:
            socat_cell = ws.cell(row=2, column=4, value='Purchase Analysis Report')

        socat_cell.font = Font(bold=True, size=10)
        socat_cell.alignment = Alignment(horizontal='center', vertical='center')

        socat_cell = ws.cell(row=3, column=4, value=f'Period : {desired_order[-2]} to {desired_order[0]}')
        socat_cell.font = Font(bold=True, size=10)
        socat_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Apply bold font to rows 5 and 6 from column G onwards
        bold_font = Font(bold=True)

        # Determine the end column (last column with data)
        end_column = ws.max_column

        # Apply bold font to rows 5 and 6 from column G onwards
        for col_idx in range(7, end_column + 1):  # Columns G onwards
            for row_idx in [5, 6]:  # Rows 5 and 6
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = bold_font

        # Initialize the row index where totals will be placed
        total_row_index = ws.max_row + 1  # Place totals in the row after the last row with data

        # Iterate through each column starting from column G
        for col_idx in range(7, end_column + 1):
            total = 0
            # Iterate through each row starting from the 7th row
            for row_idx in range(7, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                # Check if cell value is numeric before adding
                if isinstance(cell_value, (int, float)):
                    total += cell_value

            # Write the total to the cell at the designated total row and current column index
            total_cell = ws.cell(row=total_row_index, column=col_idx, value=total)
            total_cell.font = Font(bold=True)  # Bold font for total cell

        # Save the workbook
        wb.save(file_path)

        print(f"Data has been saved to {file_path} starting from the 6th row with adjusted column widths.")
        status = "success"
        message = "success"

    except Exception as error:
        print('The cause of error -->', error)
        status = 'failed'
        message = 'failed'

    return status, message, file_name, file_path
