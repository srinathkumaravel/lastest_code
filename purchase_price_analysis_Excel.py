import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from database import get_database_engine_e_eiis


def fetch_data_for_excel_report(period, excel_type):
    df_first_six = None
    df_after_six = None
    length_of_column = None
    QTY_1 = 0
    QTY_2 = 0
    QTY_3 = 0
    Amt_1 = 0
    Amt_2 = 0
    Amt_3 = 0
    AvgUnitPrice_1 = 0
    AvgUnitPrice_2 = 0
    AvgUnitPrice_3 = 0
    avg_Amt = 0
    avg_QTY = 0
    avg_UnitPrice = 0
    from_month = None
    mid_month = None
    to_month = None

    try:

        given_date = datetime.strptime(period, '%Y-%m-%d')

        # Calculate the first date of the previous month
        first_day_previous_month = (given_date - relativedelta(months=1)).replace(day=1)

        # Calculate the first date of two months ago
        first_day_two_months_ago = (given_date - relativedelta(months=2)).replace(day=1)

        # Print the results
        current_month = given_date.strftime('%Y-%m-%d')
        prev_month = first_day_previous_month.strftime('%Y-%m-%d')
        second_prev_month = first_day_two_months_ago.strftime('%Y-%m-%d')
        print('Months to be fetched data -->', current_month, ',', prev_month, ',', second_prev_month)

        # Convert the string to a datetime object
        date_obj = datetime.strptime(current_month, '%Y-%m-%d')

        # Format the datetime object to the desired format
        from_month = date_obj.strftime('%b-%Y')

        # Convert the string to a datetime object
        date_obj = datetime.strptime(prev_month, '%Y-%m-%d')

        # Format the datetime object to the desired format
        mid_month = date_obj.strftime('%b-%Y')

        # Convert the string to a datetime object
        date_obj = datetime.strptime(second_prev_month, '%Y-%m-%d')

        # Format the datetime object to the desired format
        to_month = date_obj.strftime('%b-%Y')

        # purchase_price_analysis_cash_purchase
        if excel_type == "1":
            try:
                engine = get_database_engine_e_eiis()
                # Query to fetch item categories
                cash_purchase_query = """SELECT 
                                            mc.Name AS category_name,
                                            mc.ITEM_CAT_PK AS category_id,
                                            i.ITEM_ID,
                                            i.ITEM_NAME,
                                            i.PACKAGE_ID,
                                            mi.UOM,
                                            -- Current Month
                                            ROUND(COALESCE(SUM(sd.QTY), 0), 3) AS QTY,
                                            ROUND(COALESCE(SUM(sd.QTY * sd.GP), 0), 3) AS Amt,
                                            ROUND(COALESCE((SUM(sd.QTY * sd.GP) / SUM(sd.QTY)), 0), 3) AS AvgUnitPrice,
                                            -- Previous Month
                                            ROUND(COALESCE(SUM(prev1_sd.QTY), 0), 3) AS PrevMonth_QTY,
                                            ROUND(COALESCE(SUM(prev1_sd.QTY * prev1_sd.GP), 0), 3) AS PrevMonth_Amt,
                                            ROUND(COALESCE((SUM(prev1_sd.QTY * prev1_sd.GP) / SUM(prev1_sd.QTY)), 0), 3) AS PrevMonth_AvgUnitPrice,
                                            -- Two Months Ago
                                            ROUND(COALESCE(SUM(prev2_sd.QTY), 0), 3) AS Prev2Months_QTY,
                                            ROUND(COALESCE(SUM(prev2_sd.QTY * prev2_sd.GP), 0), 3) AS Prev2Months_Amt,
                                            ROUND(COALESCE((SUM(prev2_sd.QTY * prev2_sd.GP) / SUM(prev2_sd.QTY)), 0), 3) AS Prev2Months_AvgUnitPrice
                                        FROM 
                                            suppdeldetail sd
                                        INNER JOIN 
                                            (SELECT 
                                                 GRN_ID
                                             FROM 
                                                 suppdelhead
                                             WHERE 
                                                 SUPPLIER_ID = 's00000'
                                                 AND PERIOD = %s) sh 
                                            ON sh.GRN_ID = sd.GRN_ID
                                        LEFT JOIN 
                                            suppdeldetail prev1_sd 
                                            ON prev1_sd.ITEM_ID = sd.ITEM_ID
                                            AND prev1_sd.GRN_ID IN (
                                                SELECT 
                                                    GRN_ID
                                                FROM 
                                                    suppdelhead
                                                WHERE 
                                                    SUPPLIER_ID = 's00000'
                                                    AND PERIOD = %s
                                            )
                                        LEFT JOIN 
                                            suppdeldetail prev2_sd 
                                            ON prev2_sd.ITEM_ID = sd.ITEM_ID
                                            AND prev2_sd.GRN_ID IN (
                                                SELECT 
                                                    GRN_ID
                                                FROM 
                                                    suppdelhead
                                                WHERE 
                                                    SUPPLIER_ID = 's00000'
                                                    AND PERIOD = %s
                                            )
                                        INNER JOIN 
                                            mst_item_category mc 
                                            ON LEFT(sd.ITEM_ID, 2) = mc.ITEM_CAT_PK 
                                        INNER JOIN  
                                            item i
                                            ON sd.ITEM_ID = i.ITEM_ID 
                                        INNER JOIN  
                                            mst_item_master mi
                                            ON sd.ITEM_ID = mi.ITEM_CODE     
                                        GROUP BY 
                                            mc.Name, 
                                            mc.ITEM_CAT_PK, 
                                            i.ITEM_ID, 
                                            i.ITEM_NAME, 
                                            i.PACKAGE_ID,
                                            mi.UOM
                                        ORDER BY 
                                            mc.ITEM_CAT_PK;
                                        """
                cash_purchase_df = pd.read_sql_query(cash_purchase_query, engine,
                                                     params=(current_month, prev_month, second_prev_month))
                if len(cash_purchase_df) == 0:
                    status = "success"
                    message = "No data available"
                    print("MESSAGE -->", message)
                    return df_first_six, df_after_six, length_of_column, from_month, mid_month, to_month, QTY_1, QTY_2, QTY_3, Amt_1, Amt_2, Amt_3, AvgUnitPrice_1, AvgUnitPrice_2, AvgUnitPrice_3, avg_Amt, avg_QTY, avg_UnitPrice, status, message

                # Identify columns after the first six
                columns_after_six = cash_purchase_df.columns[6:]

                # Check for NaN values and replace them with 0.00
                cash_purchase_df[columns_after_six] = cash_purchase_df[columns_after_six].fillna(0.00)
                # Split the DataFrame
                df_first_six = cash_purchase_df.iloc[:, :6]
                df_after_six = cash_purchase_df.iloc[:, 6:]

            except Exception as error:
                status = "failed"
                message = "failed"
                print('The cause of error -->', error)
                return df_first_six, df_after_six, length_of_column, from_month, mid_month, to_month, QTY_1, QTY_2, QTY_3, Amt_1, Amt_2, Amt_3, AvgUnitPrice_1, AvgUnitPrice_2, AvgUnitPrice_3, avg_Amt, avg_QTY, avg_UnitPrice, status, message

        # purchase_price_analysis_out_of_catalogue_item
        elif excel_type == "2":
            try:
                engine = get_database_engine_e_eiis()
                # Query to fetch item categories
                out_of_catalogue_item_query = """
                                                SELECT 
                                                    mc.Name AS category_name,
                                                    mc.ITEM_CAT_PK AS category_id,
                                                    mi.ITEM_CODE,
                                                    mi.ITEM_NAME,
                                                    mi.PACKAGE_ID,
                                                    mi.UOM,
                                                    -- Current Month
                                                    ROUND(COALESCE(SUM(sd.QTY), 0), 3) AS QTY,
                                                    ROUND(COALESCE(SUM(sd.QTY * sd.GP), 0), 3) AS Amt,
                                                    ROUND(COALESCE((SUM(sd.QTY * sd.GP) / SUM(sd.QTY)), 0), 3) AS AvgUnitPrice,
                                                    -- Previous Month
                                                    ROUND(COALESCE(SUM(prev1.QTY), 0), 3) AS PrevMonth_QTY,
                                                    ROUND(COALESCE(SUM(prev1.Amt), 0), 3) AS PrevMonth_Amt,
                                                    ROUND(COALESCE((SUM(prev1.Amt) / SUM(prev1.QTY)), 0), 3) AS PrevMonth_AvgUnitPrice,
                                                    -- Two Months Ago
                                                    ROUND(COALESCE(SUM(prev2.QTY), 0), 3) AS Prev2Months_QTY,
                                                    ROUND(COALESCE(SUM(prev2.Amt), 0), 3) AS Prev2Months_Amt,
                                                    ROUND(COALESCE((SUM(prev2.Amt) / SUM(prev2.QTY)), 0), 3) AS Prev2Months_AvgUnitPrice
                                                FROM 
                                                    suppdeldetail sd
                                                INNER JOIN 
                                                    suppdelhead sh 
                                                    ON sh.GRN_ID = sd.GRN_ID
                                                    AND sh.PERIOD = %s
                                                LEFT JOIN 
                                                    selectedsupplier ss 
                                                    ON ss.ITEM_ID = sd.ITEM_ID
                                                    AND ss.PERIOD = %s
                                                LEFT JOIN 
                                                    (SELECT sd.ITEM_ID, SUM(sd.QTY) AS QTY, SUM(sd.QTY * sd.GP) AS Amt
                                                     FROM suppdeldetail sd
                                                     INNER JOIN suppdelhead sh 
                                                     ON sh.GRN_ID = sd.GRN_ID
                                                     WHERE sh.PERIOD = %s 
                                                     GROUP BY sd.ITEM_ID) prev1
                                                    ON prev1.ITEM_ID = sd.ITEM_ID
                                                LEFT JOIN 
                                                    (SELECT sd.ITEM_ID, SUM(sd.QTY) AS QTY, SUM(sd.QTY * sd.GP) AS Amt
                                                     FROM suppdeldetail sd
                                                     INNER JOIN suppdelhead sh 
                                                     ON sh.GRN_ID = sd.GRN_ID
                                                     WHERE sh.PERIOD = %s
                                                     GROUP BY sd.ITEM_ID) prev2 
                                                    ON prev2.ITEM_ID = sd.ITEM_ID
                                                INNER JOIN 
                                                    mst_item_category mc 
                                                    ON LEFT(sd.ITEM_ID, 2) = mc.ITEM_CAT_PK
                                                INNER JOIN  
                                                    mst_item_master mi
                                                    ON sd.ITEM_ID = mi.ITEM_CODE
                                                WHERE 
                                                    ss.ITEM_ID IS NULL
                                                GROUP BY 
                                                    mc.Name, 
                                                    mc.ITEM_CAT_PK, 
                                                    mi.ITEM_CODE, 
                                                    mi.ITEM_NAME, 
                                                    mi.PACKAGE_ID,
                                                    mi.UOM
                                                ORDER BY 
                                                    mc.ITEM_CAT_PK;
                                                """
                out_of_catalogue_item_df = pd.read_sql_query(out_of_catalogue_item_query, engine, params=(
                    period, current_month, prev_month, second_prev_month))

                if len(out_of_catalogue_item_df) == 0:
                    status = "success"
                    message = "No data available"
                    print("MESSAGE -->", message)
                    return df_first_six, df_after_six, length_of_column, from_month, mid_month, to_month, QTY_1, QTY_2, QTY_3, Amt_1, Amt_2, Amt_3, AvgUnitPrice_1, AvgUnitPrice_2, AvgUnitPrice_3, avg_Amt, avg_QTY, avg_UnitPrice, status, message

                # Identify columns after the first six
                columns_after_six = out_of_catalogue_item_df.columns[6:]

                # Check for NaN values and replace them with 0.00
                out_of_catalogue_item_df[columns_after_six] = out_of_catalogue_item_df[columns_after_six].fillna(0.00)

                # Split the DataFrame
                df_first_six = out_of_catalogue_item_df.iloc[:, :6]
                df_after_six = out_of_catalogue_item_df.iloc[:, 6:]
            except Exception as error:
                status = "failed"
                message = "failed"
                print('The cause of error -->', error)
                return df_first_six, df_after_six, length_of_column, from_month, mid_month, to_month, QTY_1, QTY_2, QTY_3, Amt_1, Amt_2, Amt_3, AvgUnitPrice_1, AvgUnitPrice_2, AvgUnitPrice_3, avg_Amt, avg_QTY, avg_UnitPrice, status, message
                # purchase_price_analysis_out_of_catalogue_item
        else:
            try:
                engine = get_database_engine_e_eiis()
                # Query to fetch purchase price analysis
                purchase_price_analysis_query = """SELECT 
                                                        mc.Name AS category_name,
                                                        mc.ITEM_CAT_PK AS category_id,
                                                        i.ITEM_ID,
                                                        i.ITEM_NAME,
                                                        i.PACKAGE_ID,
                                                        mi.UOM,
                                                        -- Current Month
                                                        ROUND(SUM(sd.QTY), 3) AS QTY,
                                                        ROUND(SUM(sd.QTY * sd.GP), 3) AS Amt,
                                                        ROUND(SUM(sd.QTY * sd.GP) / SUM(sd.QTY), 3) AS AvgUnitPrice,
                                                        -- Previous Month
                                                        ROUND(COALESCE(SUM(prev1_sd.QTY), 0), 3) AS PrevMonth_QTY,
                                                        ROUND(COALESCE(SUM(prev1_sd.QTY * prev1_sd.GP), 0), 3) AS PrevMonth_Amt,
                                                        ROUND(COALESCE((SUM(prev1_sd.QTY * prev1_sd.GP) / SUM(prev1_sd.QTY)), 0), 3) AS PrevMonth_AvgUnitPrice,
                                                        -- Two Months Ago
                                                        ROUND(COALESCE(SUM(prev2_sd.QTY), 0), 3) AS Prev2Months_QTY,
                                                        ROUND(COALESCE(SUM(prev2_sd.QTY * prev2_sd.GP), 0), 3) AS Prev2Months_Amt,
                                                        ROUND(COALESCE((SUM(prev2_sd.QTY * prev2_sd.GP) / SUM(prev2_sd.QTY)), 0), 3) AS Prev2Months_AvgUnitPrice
                                                    FROM 
                                                        suppdeldetail sd
                                                    INNER JOIN 
                                                        (SELECT 
                                                             GRN_ID
                                                         FROM 
                                                             suppdelhead
                                                         WHERE 
                                                             PERIOD = %s) sh 
                                                        ON sh.GRN_ID = sd.GRN_ID
                                                    LEFT JOIN 
                                                        suppdeldetail prev1_sd 
                                                        ON prev1_sd.ITEM_ID = sd.ITEM_ID
                                                        AND prev1_sd.GRN_ID IN (
                                                            SELECT 
                                                                GRN_ID
                                                            FROM 
                                                                suppdelhead
                                                            WHERE 
                                                                PERIOD = %s
                                                        )
                                                    LEFT JOIN 
                                                        suppdeldetail prev2_sd 
                                                        ON prev2_sd.ITEM_ID = sd.ITEM_ID
                                                        AND prev2_sd.GRN_ID IN (
                                                            SELECT 
                                                                GRN_ID
                                                            FROM 
                                                                suppdelhead
                                                            WHERE 
                                                                PERIOD = %s
                                                        )
                                                    INNER JOIN 
                                                        mst_item_category mc 
                                                        ON LEFT(sd.ITEM_ID, 2) = mc.ITEM_CAT_PK 
                                                    INNER JOIN  
                                                        item i
                                                        ON sd.ITEM_ID = i.ITEM_ID 
                                                    INNER JOIN  
                                                        mst_item_master mi
                                                        ON sd.ITEM_ID = mi.ITEM_CODE     
                                                    GROUP BY 
                                                        mc.Name, 
                                                        mc.ITEM_CAT_PK, 
                                                        i.ITEM_ID, 
                                                        i.ITEM_NAME, 
                                                        i.PACKAGE_ID,
                                                        mi.UOM
                                                    ORDER BY 
                                                        mc.ITEM_CAT_PK;

                                                """
                purchase_price_analysis_query_df = pd.read_sql_query(purchase_price_analysis_query, engine, params=(
                    current_month, prev_month, second_prev_month))

                if len(purchase_price_analysis_query_df) == 0:
                    status = "success"
                    message = "No data available"
                    print("MESSAGE -->", message)
                    return df_first_six, df_after_six, length_of_column, from_month, mid_month, to_month, QTY_1, QTY_2, QTY_3, Amt_1, Amt_2, Amt_3, AvgUnitPrice_1, AvgUnitPrice_2, AvgUnitPrice_3, avg_Amt, avg_QTY, avg_UnitPrice, status, message

                # Identify columns after the first six
                columns_after_six = purchase_price_analysis_query_df.columns[6:]

                # Check for NaN values and replace them with 0.00
                purchase_price_analysis_query_df[columns_after_six] = purchase_price_analysis_query_df[columns_after_six].fillna(
                    0.00)

                # Split the DataFrame
                df_first_six = purchase_price_analysis_query_df.iloc[:, :6]
                df_after_six = purchase_price_analysis_query_df.iloc[:, 6:]
            except Exception as error:
                status = "failed"
                message = "failed"
                print('The cause of error -->', error)
                return df_first_six, df_after_six, length_of_column, from_month, mid_month, to_month, QTY_1, QTY_2, QTY_3, Amt_1, Amt_2, Amt_3, AvgUnitPrice_1, AvgUnitPrice_2, AvgUnitPrice_3, avg_Amt, avg_QTY, avg_UnitPrice, status, message

        rename = {'category_name': 'Category Name',
                  'category_id': 'Category ID',
                  'ITEM_CODE': 'Item Code',
                  'ITEM_NAME': 'Item Name',
                  'PACKAGE_ID': 'Package ID',
                  'UOM': 'UOM'}

        # Renaming in place
        df_first_six.rename(columns=rename, inplace=True)
        # df_after_six = df_after_six.round(3)
        # Calculate average and create new column 'avg_qty'
        df_after_six['avg_qty'] = df_after_six[['QTY', 'PrevMonth_QTY', 'Prev2Months_QTY']].mean(axis=1).round(3)
        df_after_six['avg_Amt'] = df_after_six[['Amt', 'PrevMonth_Amt', 'Prev2Months_Amt']].mean(axis=1).round(3)
        df_after_six['avg_UnitPrice'] = df_after_six[
            ['AvgUnitPrice', 'PrevMonth_AvgUnitPrice', 'Prev2Months_AvgUnitPrice']].mean(axis=1).round(3)

        QTY_1 = df_after_six['QTY'].sum()
        Amt_1 = df_after_six['Amt'].sum()
        AvgUnitPrice_1 = df_after_six['AvgUnitPrice'].sum()

        QTY_2 = df_after_six['PrevMonth_QTY'].sum()
        Amt_2 = df_after_six['PrevMonth_Amt'].sum()
        AvgUnitPrice_2 = df_after_six['PrevMonth_AvgUnitPrice'].sum()

        QTY_3 = df_after_six['Prev2Months_QTY'].sum()
        Amt_3 = df_after_six['Prev2Months_Amt'].sum()
        AvgUnitPrice_3 = df_after_six['Prev2Months_AvgUnitPrice'].sum()

        avg_QTY = df_after_six['avg_qty'].sum()
        avg_Amt = df_after_six['avg_Amt'].sum()
        avg_UnitPrice = df_after_six['avg_UnitPrice'].sum()

        rename = {'QTY': 'Quantity',
                  'Amt': 'Amount',
                  'AvgUnitPrice': 'Avg Unit Price',
                  'PrevMonth_QTY': 'Quantity',
                  'PrevMonth_Amt': 'Amount',
                  'PrevMonth_AvgUnitPrice': 'Avg Unit Price',
                  'Prev2Months_QTY': 'Quantity',
                  'Prev2Months_Amt': 'Amount',
                  'Prev2Months_AvgUnitPrice': 'Avg Unit Price',
                  'avg_qty': 'Avg Quantity',
                  'avg_Amt': 'Avg Amount',
                  'avg_UnitPrice': 'Avg Unit Price'}
        # Renaming in place
        df_after_six.rename(columns=rename, inplace=True)

        length_of_column = len(df_after_six['Avg Quantity'])
        status = "success"
        message = "success"

    except Exception as error:
        print('The cause of error -->', error)
        status = "failed"
        message = "failed"

    return df_first_six, df_after_six, length_of_column, from_month, mid_month, to_month, QTY_1, QTY_2, QTY_3, Amt_1, Amt_2, Amt_3, AvgUnitPrice_1, AvgUnitPrice_2, AvgUnitPrice_3, avg_Amt, avg_QTY, avg_UnitPrice, status, message


def create_purchase_price_analysis_excel_report(df_first_six, df_after_six, length_of_column, from_month, mid_month,
                                                to_month, QTY_1, QTY_2, QTY_3, Amt_1, Amt_2, Amt_3, AvgUnitPrice_1,
                                                AvgUnitPrice_2, AvgUnitPrice_3, avg_Amt, avg_QTY, avg_UnitPrice
                                                , excel_type):
    file_path = None
    file_name = None
    try:
        # Create a new Excel workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Convert the DataFrame to a list of lists for df_first_six
        data_first_six = [df_first_six.columns.tolist()] + df_first_six.values.tolist()

        # Convert the DataFrame to a list of lists for df_after_six
        data_after_six = [df_after_six.columns.tolist()] + df_after_six.values.tolist()

        # Write df_first_six to the Excel worksheet starting from the 6th row
        start_row = 6
        for row_num, row_data in enumerate(data_first_six, start=start_row):
            for col_num, cell_value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)

        # Write df_after_six to the Excel worksheet starting from the 6th row, column G
        for row_num, row_data in enumerate(data_after_six, start=start_row):
            for col_num, cell_value in enumerate(row_data, start=7):  # Starting from column G (index 7)
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)

        # Adjust column widths based on the maximum length of values in each column
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)  # Get the column letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as error:
                    print(error)
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Adjust width for padding and scale (optional)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Apply formatting to the header row (first row)
        header_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')  # Blue fill color
        header_font = Font(color='FFFFFF', bold=True)  # White font color

        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=start_row, column=col_num)
            cell.fill = header_fill
            cell.font = header_font

        # Apply formatting to rows 4 and 5 for all columns
        for row_num in [5]:  # Rows 4 and 5
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.fill = header_fill
                cell.font = header_font

        # Merge cells G5:I5 and add 'Apr-24'
        ws.merge_cells(start_row=5, start_column=7, end_row=5, end_column=9)  # G5:I5
        merged_cell = ws.cell(row=5, column=7, value=f'{from_month}')
        # Set the font color to black and make it bold
        black_bold_font = Font(color='000000', bold=True)  # Black font color and bold
        merged_cell.font = black_bold_font
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')  # Center align the text

        # Merge cells J5:L5 and add 'Mar-24'
        ws.merge_cells(start_row=5, start_column=10, end_row=5, end_column=12)  # J5:L5
        merged_cell = ws.cell(row=5, column=10, value=f'{mid_month}')
        black_bold_font = Font(color='000000', bold=True)  # Black font color and bold
        merged_cell.font = black_bold_font
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')  # Center align the text

        # Merge cells M5:O5 and add 'Feb-24'
        ws.merge_cells(start_row=5, start_column=13, end_row=5, end_column=15)  # M5:O5
        merged_cell = ws.cell(row=5, column=13, value=f'{to_month}')
        black_bold_font = Font(color='000000', bold=True)  # Black font color and bold
        merged_cell.font = black_bold_font

        merged_cell.alignment = Alignment(horizontal='center', vertical='center')  # Center align the text

        # Merge cells P5:R5 and add 'Avg-3 Mon'
        ws.merge_cells(start_row=5, start_column=16, end_row=5, end_column=18)  # P5:R5
        merged_cell = ws.cell(row=5, column=16, value='Avg-3 Mon')
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')  # Center align the text

        # Apply borders to all cells in the table
        border_style = Border(left=Side(border_style='thin', color='000000'),
                              right=Side(border_style='thin', color='000000'),
                              top=Side(border_style='thin', color='000000'),
                              bottom=Side(border_style='thin', color='000000'))

        # Apply borders to df_first_six table
        for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(df_first_six) - 1,
                                min_col=1, max_col=len(df_first_six.columns)):
            for cell in row:
                cell.border = border_style

        # Apply borders to df_after_six table
        for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(df_after_six) - 1,
                                min_col=7, max_col=7 + len(df_after_six.columns) - 1):
            for cell in row:
                cell.border = border_style

        # Apply black font color to cells G6, H6, I6
        black_font = Font(color='000000')  # Black font color

        for col_letter in ['G', 'H', 'I']:
            cell = ws[f"{col_letter}6"]
            cell.font = black_font

        # Apply black font color to cells G6, H6, I6
        black_font = Font(color='000000')  # Black font color

        for col_letter in ['J', 'K', 'L']:
            cell = ws[f"{col_letter}6"]
            cell.font = black_font

        # Apply black font color to cells G6, H6, I6
        black_font = Font(color='000000')  # Black font color

        for col_letter in ['M', 'N', 'O']:
            cell = ws[f"{col_letter}6"]
            cell.font = black_font

        # Apply light green fill to columns G, H, I from row 5 to the end of the table
        light_green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                                       fill_type='solid')  # Light green fill color

        for row in ws.iter_rows(min_row=5,  # Start from row 5
                                min_col=7, max_col=9,  # Columns G to I
                                max_row=start_row + len(df_after_six)):  # End at the last row of df_after_six
            for cell in row:
                cell.fill = light_green_fill

        # Apply sandal color fill to columns J, K, L from row 5 to the end of the table
        sandal_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')  # Sandal color fill

        for row in ws.iter_rows(min_row=5,  # Start from row 5
                                min_col=10, max_col=12,  # Columns J to L
                                max_row=start_row + len(data_after_six) - 1):  # End at the last row of data_after_six
            for cell in row:
                cell.fill = sandal_fill

        # Apply light grey fill to columns M, N, O from row 5 to the end of the table
        light_grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9',
                                      fill_type='solid')  # Light grey fill color

        for row in ws.iter_rows(min_row=5,  # Start from row 5
                                min_col=13, max_col=15,  # Columns M to O
                                max_row=start_row + len(data_after_six) - 1):  # End at the last row of data_after_six
            for cell in row:
                cell.fill = light_grey_fill

        # Freeze the first 6 rows and columns A to F
        ws.freeze_panes = 'G7'  # This will freeze rows 1 to 6 and columns A to F

        # Add 'SOCAT' to the first row, D column, with bold font, font size 10, and center alignment
        socat_cell = ws.cell(row=1, column=4, value='SOCAT')
        socat_cell.font = Font(bold=True, size=10)
        socat_cell.alignment = Alignment(horizontal='center', vertical='center')

        if excel_type == "2":
            socat_cell = ws.cell(row=2, column=4, value='Purchase Analysis Report (Out of Catalogue Items)')
        elif excel_type == "1":
            socat_cell = ws.cell(row=2, column=4, value='Purchase Analysis Report (Cash Purchase Items)')
        else:
            socat_cell = ws.cell(row=2, column=4, value='Purchase Analysis Report')

        socat_cell.font = Font(bold=True, size=10)
        socat_cell.alignment = Alignment(horizontal='center', vertical='center')

        socat_cell = ws.cell(row=3, column=4, value=f'Period : {from_month} to {to_month}')
        socat_cell.font = Font(bold=True, size=10)
        socat_cell.alignment = Alignment(horizontal='center', vertical='center')

        cell_value = str(length_of_column + 7)
        # Apply bold font to the cell
        ws[f'E{cell_value}'] = "TOTAL : "
        cell.font = Font(bold=True)
        ws[f'G{cell_value}'] = QTY_1
        ws[f'H{cell_value}'] = Amt_1
        ws[f'I{cell_value}'] = AvgUnitPrice_1
        ws[f'J{cell_value}'] = QTY_2
        ws[f'K{cell_value}'] = Amt_2
        ws[f'L{cell_value}'] = AvgUnitPrice_2
        ws[f'M{cell_value}'] = QTY_3
        ws[f'N{cell_value}'] = Amt_3
        ws[f'O{cell_value}'] = AvgUnitPrice_3
        ws[f'P{cell_value}'] = avg_QTY
        ws[f'Q{cell_value}'] = avg_Amt
        ws[f'R{cell_value}'] = avg_UnitPrice

        # Get the current date and time
        current_datetime = datetime.now().strftime('%Y%m%d_%H%M%S')
        if excel_type == "1":
            file_name = f"PurchaseAnalysisReport_CashPurchaseItems_{current_datetime}.xlsx"
        elif excel_type == "2":
            file_name = f"PurchaseAnalysisReport_OutOfCatalogueItems_{current_datetime}.xlsx"
        else:
            file_name = f"PurchaseAnalysisReport_{current_datetime}.xlsx"
        # Create a filename with the current date and time
        file_path = fr"C:\Users\Administrator\Downloads\eiis\purchase_analysis_excel\{file_name}"

        # Save the workbook
        wb.save(file_path)

        print(f"DataFrame written to '{file_path}' starting from the 6th row.")
        status = "success"
        message = "success"
    except Exception as error:
        print('The cause of error -->', error)
        status = "failed"
        message = "failed"

    return status, message, file_name, file_path
