import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from database import get_database_connection_e_eiis
from datetime import datetime


# Function to map the screen_code to the dictionary values
def map_screen_code(code):
    # Data dictionary for mapping
    data_dict = {
        "QR": ["QuotationRequest", "Tender Process"],
        "CQR": ["Consolidation", "Tender Process"],
        "PQ": ["PrepareQuotation", "Tender Process"],
        "QRE": ["QuotationReply", "Tender Process"],
        "PC": ["PriceComputation", "Tender Process"],
        "PCP": ["PriceComparisonPreview", "Tender Process"],
        "CSSSFI": ["ChangeSysSelecSupp", "Tender Process"],
        "ESSSFI": ["EditSysSelecSupp", "Tender Process"],
        "FTSS": ["FinalizeTheSupplierSelection", "Tender Process"],
        "DSS": ["DuplicateSupplierSelection", "Tender Process"],
        "PPC": ["PurchasePeriodClosing", "Period Closing"],
        "SPC": ["StockPeriodClosing", "Period Closing"],
        "LR": ["LocationRequest", "Location Request"],
        "LRBU": ["LocationRequestBu", "Location Request"],
        "ELR": ["EditLocationRequest", "Location Request"],
        "CTDLL": ["ChangeDeliveryLoc", "Location Request"],
        "CTDLS": ["ChangeDeliveryLocSup", "Location Request"],
        "AGPO": ["AutoPo", "Purchase Order"],
        "PO": ["ManualPoCreation", "Purchase Order"],
        "RIFS": ["ReceiveItemFromSuppl", "Stock Receive"],
        "RI": ["ReceiveInvoice", "Stock Receive"],
        "RIFL": ["ItemFromLocation", "Stock Receive"],
        "PS": ["PhysicalStock", "Stock"],
        "OCD": ["OCD", "Cash"],
        "DITL": ["DeliveryItemToLocation", "Stock Delivery"],
        "RITS": ["ReturnItemToSupplier", "Stock Delivery"],
        "RCN": ["ReceiveCreditNote", "Stock Delivery"],
        "SC": ["SupplierCreation", "Supplier"],
        "RIWS": ["RelateItemWithSupplier", "Supplier"],
        "IRM": ["ItemRelated", "Master"],
        "CM": ["Common", "Master"],
        "ARIS": ["AllReportsIncludingSaving", "Reports"],
        "ARES": ["AllReportsExcludingSaving", "Reports"]
    }

    if code in data_dict:
        return data_dict[code]
    else:
        return [None, None]


def fetch_and_create_user_rights_excel():
    try:
        with get_database_connection_e_eiis() as conn:
            # Define the parameters
            status = "A"
            excluded_types = (0, 1)

            # Create the SQL query with placeholders
            placeholders = ', '.join(['%s'] * len(excluded_types))

            cursor = conn.cursor()
            sql_query = f"""
                        SELECT user.MU_USER_PK, user.MU_USER_TYPE, user.MU_USER_NAME, screen.USER_FK, screen.SCREEN_CODE
                        FROM mst_user AS user
                        INNER JOIN mst_screen_rights_iis AS screen ON user.MU_USER_PK = screen.USER_FK
                        WHERE user.MU_STATUS = %s AND user.MU_USER_TYPE NOT IN ({placeholders});
                                            """
            # Combine the parameters
            params = (status, *excluded_types)
            cursor.execute(sql_query, params)
            data = cursor.fetchall()
            print(data)
            print(len(data))
            cursor.close()
            if len(data) == 0:
                status = "success"
                message = "No data available"
                file_name = None
                file_path = None
                return status, message, file_path, file_name
        # Convert the result to a DataFrame
        df = pd.DataFrame(data, columns=['MU_USER_PK', 'MU_USER_TYPE', 'MU_USER_NAME', 'USER_FK', 'SCREEN_CODE'])

        # Apply the mapping function
        df[['Type', 'Category']] = df['SCREEN_CODE'].apply(map_screen_code).apply(pd.Series)
        # Pivot the DataFrame
        pivot_df = df.pivot_table(index=['MU_USER_PK', 'MU_USER_TYPE', 'USER_FK', 'SCREEN_CODE', 'Type', 'Category'],
                                  columns='MU_USER_NAME', aggfunc='size', fill_value=0)

        # Flatten the multi-level columns
        pivot_df.columns = pivot_df.columns.get_level_values(0)

        # Sort columns alphabetically
        pivot_df = pivot_df.reindex(sorted(pivot_df.columns), axis=1)

        # Replace 1 with "YES" and 0 with "NO"
        pivot_df = pivot_df.replace({1: 'YES', 0: 'NO'})

        # Reset index to turn multi-index into columns
        pivot_df.reset_index(inplace=True)
        pivot_df.drop(columns=['MU_USER_PK', 'MU_USER_TYPE', 'USER_FK', 'SCREEN_CODE'], inplace=True)

        # Reorder columns
        new_column_order = ['Category', 'Type'] + [col for col in pivot_df.columns if col not in ['Category', 'Type']]
        pivot_df = pivot_df[new_column_order]

        # Define the desired order of SCREEN_CODE_HEADING values
        desired_order = ["Tender Process", "Period Closing", "Location Request", "Purchase Order", "Stock Receive",
                         "Stock", "Cash", "Stock Delivery", "Supplier", "Master", "Reports"]

        # Create a categorical type with the desired order
        pivot_df['Category'] = pd.Categorical(pivot_df['Category'], categories=desired_order, ordered=True)

        pivot_df = pivot_df.sort_values('Category').reset_index(drop=True)

        # Display the final DataFrame
        print("Final DataFrame sorted by Types and without duplicates:")
        print(pivot_df)
        status, message, file_path, file_name = create_excel_report(pivot_df)
    except Exception as error:
        print('The cause of error -->', error)
        status = "failed"
        message = "failed"
        file_name = None
        file_path = None
    return status, message, file_path, file_name


def create_excel_report(pivot_df):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Define a font for the headers (bold)
        header_font = Font(bold=True)

        # Define a border style
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Write the custom header to the worksheet
        custom_header = "OMAN\nScreen Rights"
        cell = ws.cell(row=1, column=3, value=custom_header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Merge cells for the custom header
        ws.merge_cells('C1:G4')

        # Write the headers to the worksheet starting from row 5
        for col_num, column_title in enumerate(pivot_df.columns, 1):
            cell = ws.cell(row=5, column=col_num, value=column_title)
            cell.font = header_font
            cell.border = thin_border

        # Write the DataFrame rows to the worksheet starting from row 6
        for r_idx, row in pivot_df.iterrows():
            for c_idx, value in enumerate(row):
                cell = ws.cell(row=r_idx + 6, column=c_idx + 1, value=value)
                cell.border = thin_border

        # Merge cells in the Category column based on consecutive values and center the text
        current_category = None
        start_row = 6
        for r_idx, row in pivot_df.iterrows():
            category = row['Category']
            if category != current_category:
                if current_category is not None:
                    end_row = r_idx + 5
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                    cell = ws.cell(start_row, 1)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                current_category = category
                start_row = r_idx + 6
        end_row = len(pivot_df) + 5
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        cell = ws.cell(start_row, 1)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Get the current date and time
        current_datetime = datetime.now().strftime('%Y%m%d_%H%M%S')

        file_name = f"User_rights_{current_datetime}.xlsx"
        # Create a filename with the current date and time
        file_path = fr"C:\Users\Administrator\Downloads\eiis\user_access_report\{file_name}"
        # Save the workbook
        wb.save(file_path)
        print("Excel file created successfully.")
        status = "success"
        message = "success"
    except Exception as error:
        print('The cause of error -->', error)
        status = "failed"
        message = "failed"
        file_name = None
        file_path = None

    return status, message, file_path, file_name
